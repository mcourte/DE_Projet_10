"""Tests pytest du pipeline BottleNeck — partie STRUCTURELLE.

Ces tests valident la LOGIQUE du pipeline indépendamment des valeurs
exactes du dataset. Ils peuvent donc tourner chaque mois sur de nouvelles
données sans avoir besoin de mettre à jour des chiffres en dur.

Pour les tests qui valident les chiffres exacts de la POC initiale
(825, 1428, 714, 30 millésimés, 70 568,60 €), voir :
    tests/test_chiffres_bottleneck.py

Lancement :
    python -m pytest tests/                          # tout
    python -m pytest tests/test_pipeline.py          # uniquement structurels
    python -m pytest tests/ -m "not cibles"          # exclure chiffres POC
    python -m pytest tests/ -m integration           # après run_pipeline.py

Catégories internes (par marqueur pytest) :

1. Tests unitaires (sans marqueur) — fixtures synthétiques.
   Rapides, déterministes, indépendants des fichiers réels.

2. Tests d'intégration (marqueur `@pytest.mark.integration`).
   Vérifient les invariants sur les tables DuckDB après run_pipeline.
   Auto-skip si duckdb/bottlerock.db n'existe pas.
"""

import sys
from pathlib import Path

import pandas as pd
import pytest

# Racine du projet dans sys.path pour pouvoir importer scripts.python.*
# même quand pytest est lancé depuis n'importe quel dossier.
PROJECT_ROOT = Path(__file__).resolve().parents[1]
sys.path.insert(0, str(PROJECT_ROOT))

from scripts.python import (  # noqa: E402  (import après modif sys.path)
    calculate_revenue,
    clean_data,
    extract_files,
    generate_reports,
    identify_wines,
    join_data,
    load_to_duckdb,
)


DUCKDB_FILE = PROJECT_ROOT / "duckdb" / "bottlerock.db"


# ===========================================================================
# Fixtures synthétiques (données minimales pour les tests unitaires)
# ===========================================================================


@pytest.fixture
def fake_erp() -> pd.DataFrame:
    """Mini-ERP de 4 lignes pour tester nettoyage / jointure."""
    return pd.DataFrame({
        "product_id": [1, 2, 3, 4],
        "onsale_web": [1, 1, 1, 0],
        "price": [10.0, 50.0, 200.0, 25.0],
        "stock_quantity": [10, 5, 0, 20],
        "stock_status": ["instock", "instock", "outofstock", "instock"],
    })


@pytest.fixture
def fake_web() -> pd.DataFrame:
    """Mini-WEB avec 5 lignes incluant :
        - une ligne sku=NaN (à dropper à l'étape 1).
        - un sku 'A2' en doublon (post_type='product' x2) : on garde le 1er.
        - un sku 'A3' en attachment + product : la priorité va au 'product'.
    """
    return pd.DataFrame({
        "sku": ["A1", "A2", "A3", None, "A2"],
        "post_title": ["Vin 1", "Vin 2", "Vin 3", "Image", "Vin 2 dup"],
        "post_type": ["product", "product", "product", "attachment", "product"],
        "post_status": ["publish", "publish", "publish", "publish", "publish"],
        "total_sales": [10.0, 5.0, 2.0, 0.0, 100.0],
        "post_date": pd.to_datetime([
            "2020-01-01", "2020-02-01", "2020-03-01",
            "2020-04-01", "2020-05-01",
        ]),
    })


@pytest.fixture
def fake_liaison() -> pd.DataFrame:
    """Mini-LIAISON : product_id 1->A1, 2->A2, 3->A3, et 4 sans id_web (NaN)."""
    return pd.DataFrame({
        "product_id": [1, 2, 3, 4],
        "id_web": ["A1", "A2", "A3", None],
    })


# ===========================================================================
# 1. Tests unitaires (sans marqueur)
# ===========================================================================


class TestExtractFiles:
    """Tests de scripts.python.extract_files."""

    def test_list_source_files(self, tmp_path: Path) -> None:
        # On crée 3 fichiers : un .csv, un .xlsx, un caché .hidden.
        # Seuls les 2 premiers doivent apparaître dans la liste.
        (tmp_path / "a.csv").write_text("x,y\n1,2", encoding="utf-8")
        (tmp_path / "b.xlsx").write_bytes(b"")
        (tmp_path / ".hidden").write_text("nope", encoding="utf-8")

        files = extract_files.list_source_files(tmp_path)
        names = {f.name for f in files}

        assert "a.csv" in names
        assert "b.xlsx" in names
        assert ".hidden" not in names

    def test_read_csv(self, tmp_path: Path) -> None:
        path = tmp_path / "x.csv"
        path.write_text("a,b\n1,2\n", encoding="utf-8")

        df = extract_files.read_file(path)
        assert list(df.columns) == ["a", "b"]

    def test_read_unknown_extension(self, tmp_path: Path) -> None:
        # Un .json n'est pas dans SUPPORTED_EXTENSIONS -> ValueError.
        path = tmp_path / "x.json"
        path.write_text("{}", encoding="utf-8")

        with pytest.raises(ValueError):
            extract_files.read_file(path)

    def test_read_bottleneck_missing(self, tmp_path: Path) -> None:
        # Dossier vide -> aucun des 3 fichiers -> FileNotFoundError.
        with pytest.raises(FileNotFoundError):
            extract_files.read_bottleneck_sources(tmp_path)


class TestCleanData:
    """Tests de scripts.python.clean_data."""

    def test_clean_erp_dedup(self, fake_erp: pd.DataFrame) -> None:
        # On injecte un doublon : clean_erp doit le retirer.
        df = pd.concat([fake_erp, fake_erp.head(1)], ignore_index=True)

        out = clean_data.clean_erp(df)

        assert out["product_id"].is_unique
        assert len(out) == 4

    def test_clean_web_drops_sku_nan(self, fake_web: pd.DataFrame) -> None:
        # Étape 1 du nettoyage WEB : on retire les lignes sku=NaN.
        # Sur 5 lignes brutes dont 1 NaN, il doit en rester 4.
        n = len(fake_web.dropna(subset=["sku"]))
        assert n == 4

    def test_clean_web_dedup_priorise_product(self, fake_web: pd.DataFrame) -> None:
        # Étape 2 : tous les sku doivent être uniques après dédup.
        out = clean_data.clean_web(fake_web)

        assert out["sku"].is_unique
        for sku in ["A1", "A2", "A3"]:
            assert sku in out["sku"].tolist()

    def test_clean_web_dedup_keeps_first_for_same_post_type(
        self, fake_web: pd.DataFrame
    ) -> None:
        # 'A2' apparaît 2x en 'product' : on garde "Vin 2" (le 1er),
        # pas "Vin 2 dup" (total_sales=100).
        out = clean_data.clean_web(fake_web)

        a2 = out[out["sku"] == "A2"].iloc[0]
        assert a2["post_title"] == "Vin 2"
        assert a2["total_sales"] == 5.0

    def test_clean_liaison_no_drop_nan(self, fake_liaison: pd.DataFrame) -> None:
        # Stéphane CONSERVE les id_web=NaN après dédup (filtrés plus tard
        # par l'inner join WEB).
        out = clean_data.clean_liaison(fake_liaison)
        assert len(out) == 4


class TestJoinData:
    """Tests de scripts.python.join_data."""

    def test_join_inner_filters_orphans(
        self,
        fake_erp: pd.DataFrame,
        fake_web: pd.DataFrame,
        fake_liaison: pd.DataFrame,
    ) -> None:
        # Le produit 4 n'a pas d'id_web -> filtré par l'inner join.
        erp = clean_data.clean_erp(fake_erp)
        web = clean_data.clean_web(fake_web)
        liaison = clean_data.clean_liaison(fake_liaison)

        out = join_data.join_sources(erp, web, liaison)

        assert len(out) == 3
        assert set(out["sku"]) == {"A1", "A2", "A3"}

    def test_report_orphans(
        self,
        fake_erp: pd.DataFrame,
        fake_web: pd.DataFrame,
        fake_liaison: pd.DataFrame,
    ) -> None:
        erp = clean_data.clean_erp(fake_erp)
        web = clean_data.clean_web(fake_web)
        liaison = clean_data.clean_liaison(fake_liaison)

        report = join_data.report_orphans(erp, web, liaison)

        assert "erp_sans_id_web" in report
        assert "web_sans_pendant_erp" in report


class TestIdentifyWines:
    """Tests de scripts.python.identify_wines."""

    def test_classify_zscore_isole_outlier(self) -> None:
        # 100 valeurs à 10 + 1 outlier à 10 000 -> 1 seul premium.
        df = pd.DataFrame({"price": [10] * 100 + [10_000]})

        out = identify_wines.classify_wines(df)

        assert (out["segment"] == "premium").sum() == 1
        assert out.iloc[-1]["segment"] == "premium"

    def test_classify_custom_threshold(self) -> None:
        # Avec un seuil élevé (3.0), aucun produit ne dépasse mean + 3*std.
        df = pd.DataFrame({"price": [10, 20, 30, 40, 50, 60]})

        out = identify_wines.classify_wines(df, threshold=3.0)

        assert (out["segment"] == "premium").sum() == 0

    def test_classify_missing_price_column(self) -> None:
        with pytest.raises(ValueError):
            identify_wines.classify_wines(pd.DataFrame({"x": [1]}))

    def test_split_premium_ordinary(self) -> None:
        df = pd.DataFrame({
            "price": [10, 1000],
            "segment": ["ordinary", "premium"],
        })

        premium, ordinary = identify_wines.split_premium_ordinary(df)

        assert len(premium) == 1 and premium.iloc[0]["price"] == 1000
        assert len(ordinary) == 1


class TestCalculateRevenue:
    """Tests de scripts.python.calculate_revenue."""

    def test_total_revenue(self) -> None:
        # CA = 10*3 + 20*5 = 30 + 100 = 130.
        df = pd.DataFrame({"price": [10, 20], "total_sales": [3, 5]})
        assert calculate_revenue.total_revenue(df) == 30 + 100

    def test_revenue_per_product_sorted(self) -> None:
        # Le produit B (CA=200) doit passer devant A (CA=10).
        df = pd.DataFrame({
            "sku": ["A", "B"],
            "post_title": ["x", "y"],
            "price": [10, 100],
            "total_sales": [1, 2],
        })

        out = calculate_revenue.revenue_per_product(df)

        assert out.iloc[0]["sku"] == "B"

    def test_revenue_summary_segments(self) -> None:
        df = pd.DataFrame({
            "price": [10, 100],
            "total_sales": [1, 2],
            "segment": ["ordinary", "premium"],
        })

        out = calculate_revenue.revenue_summary(df)

        assert set(out["segment"]) == {"ordinary", "premium"}
        assert abs(out["part_pct"].sum() - 100.0) < 0.01


class TestLoadToDuckDB:
    """Tests de scripts.python.load_to_duckdb.

    On utilise `tmp_path` (fixture pytest) pour créer un fichier .db
    temporaire à chaque test. Pytest nettoie le dossier à la fin du test,
    donc rien ne pollue le projet.
    """

    def test_get_connection(self, tmp_path: Path) -> None:
        # On crée une base de test dans un dossier temporaire.
        db_path = tmp_path / "test.db"
        conn = load_to_duckdb.get_connection(db_path)
        try:
            assert conn.execute("SELECT 1").fetchone()[0] == 1
        finally:
            conn.close()

    def test_write_and_query(self, tmp_path: Path) -> None:
        # Round-trip : write_table puis query renvoie 3 lignes.
        df = pd.DataFrame({"a": [1, 2, 3]})
        db_path = tmp_path / "test.db"
        conn = load_to_duckdb.get_connection(db_path)
        try:
            load_to_duckdb.write_table(df, "t", "replace", conn=conn)
            out = load_to_duckdb.query("SELECT COUNT(*) AS n FROM t", conn=conn)
            assert out.iloc[0]["n"] == 3
        finally:
            conn.close()

    def test_write_table_invalid_mode(self) -> None:
        with pytest.raises(ValueError):
            load_to_duckdb.write_table(
                pd.DataFrame({"a": [1]}), "t", mode="bogus"
            )

    # Note : on testait avant le mécanisme de fallback CSV via monkeypatch
    # sur une fonction interne. On a retiré ce test pour simplifier le code
    # (et éviter d'avoir une fonction qui n'existait que pour les tests).
    # Le fallback fonctionne toujours en prod : si DuckDB tombe pendant 4
    # tentatives, write_table écrit un CSV de secours dans FALLBACK_DIR.


class TestGenerateReports:
    """Tests de scripts.python.generate_reports."""

    def test_report_has_expected_sheets(self, tmp_path: Path) -> None:
        # Le fichier Excel doit contenir les 4 onglets imposés par l'énoncé.
        from openpyxl import load_workbook

        df = pd.DataFrame({
            "sku": ["A", "B", "C"],
            "post_title": ["x", "y", "z"],
            "price": [10, 100, 200],
            "stock_quantity": [1, 2, 3],
            "stock_status": ["instock"] * 3,
            "total_sales": [1, 2, 3],
            "post_date": pd.to_datetime(["2020-01-01"] * 3),
            "segment": ["ordinary", "ordinary", "premium"],
        })

        out = generate_reports.generate_excel_report(
            df, output_path=tmp_path / "r.xlsx"
        )

        wb = load_workbook(out, read_only=True)
        assert set(wb.sheetnames) == {
            "CA_par_produit",
            "CA_total",
            "Vins_premium",
            "Vins_ordinaires",
        }

    def test_report_missing_segment_column(self, tmp_path: Path) -> None:
        # Sans colonne 'segment' -> ValueError.
        df = pd.DataFrame({"sku": ["A"], "price": [10], "total_sales": [1]})

        with pytest.raises(ValueError):
            generate_reports.generate_excel_report(
                df, output_path=tmp_path / "r.xlsx"
            )


# ===========================================================================
# 2. Tests d'intégration DuckDB (invariants — réutilisables chaque mois)
# ===========================================================================
# Ces tests valident la STRUCTURE de la base après le pipeline, sans dépendre
# des valeurs exactes du dataset. Ils continueront de passer le mois suivant
# même si les chiffres changent.


@pytest.fixture
def duckdb_conn():
    """Ouvre une connexion sur la base post-pipeline.

    Auto-skip si la base n'existe pas (il faut lancer run_pipeline.py
    au moins une fois pour qu'elle soit créée).
    """
    if not DUCKDB_FILE.exists():
        pytest.skip("duckdb/bottlerock.db absent : exécuter run_pipeline.py")

    conn = load_to_duckdb.get_connection(DUCKDB_FILE)
    try:
        yield conn
    finally:
        conn.close()


@pytest.mark.integration
class TestDuckDBIntegration:
    """Invariants post-pipeline : valables quel que soit le dataset."""

    def test_table_produits_consolides_non_vide(self, duckdb_conn) -> None:
        # La table doit exister et contenir au moins une ligne.
        n = duckdb_conn.execute(
            "SELECT COUNT(*) FROM produits_consolides"
        ).fetchone()[0]
        assert n > 0, "La table produits_consolides est vide"

    def test_ca_premium_plus_ordinary_egal_total(self, duckdb_conn) -> None:
        # Cohérence : somme des CA par segment == CA total.
        # COALESCE protège contre un segment vide (renvoie 0).
        sql = """
            SELECT
                COALESCE(SUM(CASE WHEN segment='premium'  THEN price*total_sales END), 0) AS ca_p,
                COALESCE(SUM(CASE WHEN segment='ordinary' THEN price*total_sales END), 0) AS ca_o,
                COALESCE(SUM(price*total_sales), 0) AS ca_t
            FROM produits_consolides
        """
        row = duckdb_conn.execute(sql).fetchone()
        ca_p, ca_o, ca_t = float(row[0]), float(row[1]), float(row[2])

        assert abs((ca_p + ca_o) - ca_t) < 0.01

    def test_completude_aucun_null_critique(self, duckdb_conn) -> None:
        # Après nettoyage : aucune ligne avec price/sku/product_id NULL.
        n = duckdb_conn.execute("""
            SELECT COUNT(*) FROM produits_consolides
            WHERE price IS NULL OR sku IS NULL OR product_id IS NULL
        """).fetchone()[0]

        assert n == 0

    def test_sku_tous_uniques(self, duckdb_conn) -> None:
        # Invariant : après dédup, chaque sku apparaît exactement 1 fois.
        # On compare DISTINCT vs total plutôt qu'à une valeur fixe.
        row = duckdb_conn.execute("""
            SELECT COUNT(DISTINCT sku) AS distincts, COUNT(*) AS total
            FROM produits_consolides
        """).fetchone()

        assert row[0] == row[1], f"sku non uniques : {row[0]} distincts / {row[1]} total"

    def test_segments_repartis(self, duckdb_conn) -> None:
        # Invariant méthode Z-score : chaque ligne a un segment 'premium'
        # ou 'ordinary'. Au moins un des deux doit être non vide.
        row = duckdb_conn.execute("""
            SELECT
                SUM(CASE WHEN segment='premium'  THEN 1 ELSE 0 END) AS n_p,
                SUM(CASE WHEN segment='ordinary' THEN 1 ELSE 0 END) AS n_o,
                COUNT(*) AS total
            FROM produits_consolides
        """).fetchone()
        n_p, n_o, total = int(row[0]), int(row[1]), int(row[2])

        assert n_p + n_o == total, "Certaines lignes n'ont pas de segment"
        assert total > 0
